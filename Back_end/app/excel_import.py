"""Excel 导入：解析 + 字段映射（规则兜底）+ 逐行规范化。

设计要点：
- 解析 Excel 全部行（受 IMPORT_MAX_ROWS 限制），但 LLM 只看表头 + 前 5 行样本；
- 逐行规范化在本地完成（日期/金额/状态/必填校验），与 LLM 无关，零额外成本；
- LLM 调用失败时自动 fallback 到内置的"中英文别名表"做规则映射。
- 自动检测真实表头行：跳过顶部的标题、说明、状态图例、空行等，
  也支持调用方通过 header_row（1-based）强制指定。
"""
from __future__ import annotations

import csv
import io
import logging
import re
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

from .asset_code import ASSET_CLASSES
from .config import get_settings
from .llm_client import (
    ALLOWED_FIELD_KEYS,
    LLMUnavailable,
    SYSTEM_FIELD_DICT,
    fill_missing_fields as llm_fill_missing_fields,
    is_configured as llm_is_configured,
    map_columns as llm_map_columns,
)

logger = logging.getLogger(__name__)


# ============ 规则兜底：中英文表头别名 → 系统字段 key ============
_ALIAS_MAP: Dict[str, str] = {}


def _add_aliases(field: str, aliases: List[str]) -> None:
    for a in aliases:
        _ALIAS_MAP[a.strip().lower()] = field


_add_aliases(
    "asset_code",
    ["资产编号", "资产编码", "编号", "编码", "资产id", "asset_code", "asset code", "assetcode"],
)
_add_aliases(
    "asset_class",
    ["资产大类", "大类", "类别代码", "asset_class", "asset class", "assetclass"],
)
_add_aliases(
    "category",
    ["资产分类", "分类", "类别", "类型", "种类", "category", "type", "kind"],
)
_add_aliases(
    "brand",
    ["品牌", "厂商", "厂牌", "生产厂家", "brand", "manufacturer", "vendor"],
)
_add_aliases(
    "model",
    ["型号", "产品型号", "model", "model no", "model number"],
)
_add_aliases(
    "serial_number",
    ["sn", "s/n", "序列号", "序号", "出厂编号", "serial", "serial number", "serialno", "serial_number"],
)
_add_aliases(
    "specification",
    ["规格", "规格配置", "配置", "参数", "specification", "spec", "specs", "config"],
)
_add_aliases(
    "purchase_date",
    ["购置日期", "采购日期", "购入日期", "入账日期", "购买日期", "入库日期",
     "purchase date", "purchase_date", "buy date"],
)
_add_aliases(
    "supplier",
    ["供应商", "采购渠道", "厂家", "购买渠道", "卖家", "supplier", "vendor"],
)
_add_aliases(
    "price",
    ["金额", "采购金额", "采购价格", "价格", "单价", "原值", "原价",
     "price", "amount", "cost"],
)
_add_aliases(
    "warranty_until",
    ["保修至", "保修到期", "保修截止", "保修日期", "质保到期",
     "warranty", "warranty until", "warranty_until"],
)
_add_aliases(
    "location",
    ["位置", "存放位置", "存放地点", "地点", "存放", "所在地", "location", "place"],
)
_add_aliases(
    "owner",
    ["使用人", "责任人", "持有人", "领用人", "负责人", "owner", "user", "holder"],
)
_add_aliases(
    "department",
    ["部门", "所属部门", "归属部门", "使用部门", "department", "dept"],
)
_add_aliases(
    "status",
    ["状态", "资产状态", "使用状态", "status", "state"],
)
_add_aliases(
    "remark",
    ["备注", "说明", "注释", "备注说明", "remark", "remarks", "note", "notes", "comment"],
)


# 资产大类 code/name 双向索引（小写）
_CLASS_CODE_SET = {c["code"].lower() for c in ASSET_CLASSES}
_CLASS_NAME_TO_CODE: Dict[str, str] = {
    c["name"].lower(): c["code"] for c in ASSET_CLASSES
}
# 一些常见同义词
_CLASS_NAME_TO_CODE.update(
    {
        "信息化": "IT",
        "信息设备": "IT",
        "电子设备": "IT",
        "办公设备": "OF",
        "办公家具": "OF",
        "家具": "OF",
        "汽车": "VH",
        "车": "VH",
        "机械": "ME",
        "设备": "ME",
        "仪器": "IN",
        "仪表": "IN",
        "其它": "OT",
    }
)


# ============ 公共工具 ============

def _normalize_header(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip()


def rule_based_mapping(headers: List[str]) -> Dict[str, Optional[str]]:
    """基于内置别名表的列映射，作为 LLM 失败时的兜底。"""
    result: Dict[str, Optional[str]] = {}
    for h in headers:
        key = (h or "").strip().lower()
        # 去掉常见无用字符
        key_clean = re.sub(r"[\s_\-/\\()\[\]（）]+", "", key)
        mapped = _ALIAS_MAP.get(key)
        if not mapped:
            for alias, field in _ALIAS_MAP.items():
                alias_clean = re.sub(r"[\s_\-/\\()\[\]（）]+", "", alias)
                if alias_clean and (alias_clean == key_clean or alias_clean in key_clean):
                    mapped = field
                    break
        result[h] = mapped
    return result


# ============ Excel / CSV 解析 ============

# 顶部最多扫描多少行用于"找真正的表头行"
_HEADER_SCAN_LIMIT = 20


def parse_workbook(
    file_bytes: bytes,
    filename: str,
    header_row: Optional[int] = None,
) -> Tuple[List[str], List[Dict[str, Any]], int]:
    """解析 Excel 或 CSV 字节流。

    参数：
        file_bytes：文件内容
        filename：用于判别后缀
        header_row：1-based 行号；为 None 则自动检测

    返回：(headers, rows, header_row_1based)
        rows 是 list of dict，键为表头列名（保持表头出现顺序）；
        header_row_1based 是实际使用的表头所在行（1-based，便于 UI 展示）。
    """
    settings = get_settings()
    max_rows = max(1, int(settings.import_max_rows))
    name = (filename or "").lower()

    if name.endswith(".csv"):
        all_rows = _read_csv_rows(file_bytes, max_rows + _HEADER_SCAN_LIMIT)
    else:
        all_rows = _read_xlsx_rows(file_bytes, max_rows + _HEADER_SCAN_LIMIT)

    if not all_rows:
        return [], [], 1

    header_idx = _resolve_header_index(all_rows, header_row)
    raw_header = all_rows[header_idx] if header_idx < len(all_rows) else []
    headers = [_normalize_header(c) for c in raw_header]
    # 去掉表头末尾连续的空列
    while headers and headers[-1] == "":
        headers.pop()
    headers = _dedup_headers(headers)

    if not headers:
        return [], [], header_idx + 1

    rows: List[Dict[str, Any]] = []
    for r in all_rows[header_idx + 1: header_idx + 1 + max_rows]:
        d: Dict[str, Any] = {}
        for j, h in enumerate(headers):
            d[h] = r[j] if j < len(r) else None
        if _row_has_any_value(d):
            rows.append(d)

    return headers, rows, header_idx + 1


def _read_csv_rows(file_bytes: bytes, max_total: int) -> List[List[Any]]:
    text: Optional[str] = None
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030", "latin-1"):
        try:
            text = file_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("无法识别 CSV 文件编码")

    reader = csv.reader(io.StringIO(text))
    rows: List[List[Any]] = []
    for i, row in enumerate(reader):
        rows.append(list(row))
        if i + 1 >= max_total:
            break
    return rows


def _read_xlsx_rows(file_bytes: bytes, max_total: int) -> List[List[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise RuntimeError(f"openpyxl 未安装：{e}") from e

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        raise ValueError(f"Excel 文件无法打开：{e}") from e

    ws = wb.active
    if ws is None:
        return []

    rows: List[List[Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append(list(row))
        if i + 1 >= max_total:
            break
    return rows


def _resolve_header_index(
    all_rows: List[List[Any]], user_header_row: Optional[int]
) -> int:
    """决定表头所在的 0-based 行索引。

    优先级：用户强制指定 > 启发式检测 > 第 0 行。
    """
    if user_header_row is not None:
        idx = int(user_header_row) - 1
        return max(0, min(idx, len(all_rows) - 1))
    return _detect_header_row(all_rows)


def _detect_header_row(all_rows: List[List[Any]]) -> int:
    """启发式：在前 _HEADER_SCAN_LIMIT 行中找最像表头的一行。

    判定特征（按重要性）：
    1. 非空单元格数量较多（跳过"只有 1 个跨列标题"那种行）；
    2. 非空单元格几乎全是字符串（数字主导多半是数据行）；
    3. 字符串普遍较短（不像"资产状态：在用|闲置|维修中|已报废"这种长描述行）；
    4. 各值彼此不重复；
    5. 表头下方应紧跟非空数量相近的数据行。

    返回 0-based 行索引；找不到时回退到 0。
    """
    scan = all_rows[:_HEADER_SCAN_LIMIT]
    if not scan:
        return 0

    counts = [_non_empty_count(r) for r in scan]
    max_count = max(counts) if counts else 0
    if max_count <= 1:
        return 0

    threshold = max(2, int(max_count * 0.6))

    best_idx = -1
    best_score = -1.0
    for i, row in enumerate(scan):
        ne = counts[i]
        if ne < threshold:
            continue

        cells = [v for v in row if _cell_is_nonempty(v)]
        strings = [v for v in cells if isinstance(v, str)]
        # 字符串占比
        str_ratio = len(strings) / len(cells) if cells else 0.0
        if str_ratio < 0.7:
            continue

        # 平均字符串长度（过长的多半是说明行而非表头）
        avg_len = sum(len(s.strip()) for s in strings) / max(1, len(strings))
        if avg_len > 30:
            continue

        # 去重后比例
        uniq_ratio = len({str(v).strip() for v in cells}) / len(cells) if cells else 0.0

        # 下一行的非空数（数据行应与表头列数相近）
        next_ne = counts[i + 1] if i + 1 < len(counts) else 0
        next_match = 1.0 if next_ne >= ne * 0.4 else 0.3

        score = (
            ne / max_count * 1.2
            + str_ratio * 0.8
            + uniq_ratio * 0.6
            + next_match * 0.6
            - min(avg_len, 30) / 60.0
        )
        if score > best_score:
            best_score = score
            best_idx = i

    return best_idx if best_idx >= 0 else 0


def _cell_is_nonempty(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, str):
        return bool(v.strip())
    return True


def _non_empty_count(row: List[Any]) -> int:
    return sum(1 for v in row if _cell_is_nonempty(v))


def _dedup_headers(headers: List[str]) -> List[str]:
    """处理空表头与重名表头，保证键唯一。"""
    seen: Dict[str, int] = {}
    result: List[str] = []
    for i, h in enumerate(headers):
        name = h or f"列{i + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        result.append(name)
    return result


def _row_has_any_value(d: Dict[str, Any]) -> bool:
    for v in d.values():
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return True
    return False


# ============ 字段映射主入口（含 LLM + 规则兜底） ============

def resolve_mapping(
    headers: List[str], samples: List[Dict[str, Any]]
) -> Tuple[Dict[str, Optional[str]], bool]:
    """返回 (mapping, llm_used)。

    - 若 KEY 已配置，优先调用 LLM；LLM 失败回退规则；
    - 若两者都没识别出某列，则规则结果中该列保持 None。
    """
    rule_mapping = rule_based_mapping(headers)
    if not llm_is_configured():
        return rule_mapping, False

    try:
        ai_mapping = llm_map_columns(headers, samples)
    except LLMUnavailable as e:
        logger.info("LLM 不可用，回退规则映射：%s", e)
        return rule_mapping, False
    except Exception as e:  # noqa: BLE001
        # 任何未预期的异常（含历史遗留的 openai/httpx proxies TypeError 等）
        # 都不应阻断 Excel 导入流程，统一回退到规则映射。
        logger.warning(
            "LLM 调用抛出未预期异常，回退规则映射：%s: %s",
            type(e).__name__,
            e,
        )
        return rule_mapping, False

    # 合并：LLM 命中优先；LLM 没命中的列用规则结果补
    merged: Dict[str, Optional[str]] = {}
    for h in headers:
        merged[h] = ai_mapping.get(h) or rule_mapping.get(h) or None
    return merged, True


# ============ 单元格值规范化 ============

_STATUS_MAP: Dict[str, str] = {
    "在用": "在用", "使用中": "在用", "使用": "在用", "正常": "在用", "在岗": "在用",
    "in use": "在用", "active": "在用", "using": "在用",
    "闲置": "闲置", "空闲": "闲置", "库存": "闲置", "未使用": "闲置",
    "idle": "闲置", "spare": "闲置",
    "维修": "维修", "维修中": "维修", "维护": "维修", "保修": "维修",
    "repair": "维修", "repairing": "维修", "maintenance": "维修",
    "报废": "报废", "已报废": "报废", "废弃": "报废", "损坏": "报废", "坏": "报废",
    "scrap": "报废", "scrapped": "报废", "broken": "报废", "disposed": "报废",
}


def _coerce_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, float) and v != v:  # NaN
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return str(v).strip() or None


def _coerce_date(v: Any) -> Tuple[Optional[str], Optional[str]]:
    """统一日期格式为 ISO 字符串 'YYYY-MM-DDTHH:MM:SS'。

    返回 (iso_str_or_None, error_message_or_None)。
    """
    if v is None:
        return None, None
    if isinstance(v, datetime):
        return v.replace(microsecond=0).isoformat(), None
    if isinstance(v, date):
        return datetime.combine(v, datetime.min.time()).isoformat(), None

    if isinstance(v, (int, float)):
        # Excel 序列日期：1900-01-01 起算
        try:
            # 兼容 openpyxl data_only=True 已经能直接给 datetime，这里只兜底数字
            base = datetime(1899, 12, 30)
            dt = base.fromordinal(int(base.toordinal() + int(v)))
            return dt.isoformat(), None
        except Exception:  # noqa: BLE001
            return None, f"无法识别的日期数值：{v}"

    s = str(v).strip()
    if not s:
        return None, None

    # 去掉中文"年月日"
    cleaned = (
        s.replace("年", "-").replace("月", "-").replace("日", "")
        .replace("/", "-").replace(".", "-")
    )
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" -")
    parts = cleaned.split(" ", 1)
    date_part = parts[0]

    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            dt = datetime.strptime(cleaned, fmt) if " " in cleaned else datetime.strptime(date_part, "%Y-%m-%d")
            return dt.isoformat(), None
        except ValueError:
            continue

    # 单独尝试日期部分
    try:
        dt = datetime.strptime(date_part, "%Y-%m-%d")
        return dt.isoformat(), None
    except ValueError:
        pass

    return None, f"无法识别的日期：{s}"


def _coerce_price(v: Any) -> Tuple[Optional[float], Optional[str]]:
    if v is None:
        return None, None
    if isinstance(v, (int, float)):
        if v < 0:
            return None, f"金额不能为负：{v}"
        return float(v), None
    s = str(v).strip()
    if not s:
        return None, None
    # 去掉常见货币符号/千分位/单位
    cleaned = re.sub(r"[¥￥$,\s]", "", s)
    cleaned = cleaned.replace("元", "").replace("RMB", "").replace("rmb", "")
    try:
        f = float(cleaned)
        if f < 0:
            return None, f"金额不能为负：{s}"
        return f, None
    except ValueError:
        return None, f"无法识别的金额：{s}"


def _coerce_status(v: Any) -> Tuple[Optional[str], Optional[str]]:
    s = _coerce_str(v)
    if s is None:
        return None, None
    mapped = _STATUS_MAP.get(s) or _STATUS_MAP.get(s.lower())
    if mapped:
        return mapped, None
    # 未识别：保留原值，但报 warning
    return s, f"未识别的状态值「{s}」，已原样保留"


def _coerce_asset_class(v: Any) -> Tuple[Optional[str], Optional[str]]:
    s = _coerce_str(v)
    if s is None:
        return None, None
    key = s.strip().lower()
    if key in _CLASS_CODE_SET:
        return s.upper(), None
    if key in _CLASS_NAME_TO_CODE:
        return _CLASS_NAME_TO_CODE[key], None
    # 部分匹配（如"信息化设备"含"信息"）
    for name, code in _CLASS_NAME_TO_CODE.items():
        if name and name in key:
            return code, f"按关键词「{name}」推断大类为 {code}"
    return None, f"无法识别的资产大类「{s}」"


# 字段处理路由表
_FIELD_HANDLERS: Dict[str, str] = {
    "purchase_date": "date",
    "warranty_until": "date",
    "price": "price",
    "status": "status",
    "asset_class": "asset_class",
}


def normalize_row(
    row_raw: Dict[str, Any], mapping: Dict[str, Optional[str]]
) -> Tuple[Dict[str, Any], List[Dict[str, str]]]:
    """把一行原始字典按 mapping 转成 AssetCreate 兼容字典 + issue 列表。

    issue 结构：{level: "error"|"warning", field: str|None, message: str}
    """
    issues: List[Dict[str, str]] = []
    data: Dict[str, Any] = {}

    # 先按 mapping 重组（同一目标字段被多列映射时，后者覆盖前者）
    for src_col, target in mapping.items():
        if not target or target not in ALLOWED_FIELD_KEYS:
            continue
        raw_val = row_raw.get(src_col)
        handler = _FIELD_HANDLERS.get(target)

        if handler == "date":
            iso, err = _coerce_date(raw_val)
            if err:
                issues.append({"level": "warning", "field": target, "message": err})
            data[target] = iso
        elif handler == "price":
            f, err = _coerce_price(raw_val)
            if err:
                issues.append({"level": "warning", "field": target, "message": err})
            data[target] = f
        elif handler == "status":
            s, warn = _coerce_status(raw_val)
            if warn:
                issues.append({"level": "warning", "field": target, "message": warn})
            data[target] = s
        elif handler == "asset_class":
            code, msg = _coerce_asset_class(raw_val)
            if msg and code is None:
                issues.append({"level": "error", "field": target, "message": msg})
            elif msg:
                issues.append({"level": "warning", "field": target, "message": msg})
            data[target] = code
        else:
            data[target] = _coerce_str(raw_val)

    # 必填校验：asset_class、category 必须有；其他缺失只是 warning
    if not data.get("asset_class"):
        issues.append(
            {"level": "error", "field": "asset_class", "message": "缺少资产大类"}
        )
    if not data.get("category"):
        issues.append(
            {"level": "error", "field": "category", "message": "缺少资产分类"}
        )

    # 默认状态
    if not data.get("status"):
        data["status"] = "在用"

    # 截断超长字符串，按 schema 的 max_length 约束（保守处理）
    _truncate_field(data, "asset_code", 64, issues)
    _truncate_field(data, "asset_class", 8, issues)
    _truncate_field(data, "category", 64, issues)
    _truncate_field(data, "brand", 64, issues)
    _truncate_field(data, "model", 128, issues)
    _truncate_field(data, "serial_number", 64, issues)
    _truncate_field(data, "specification", 255, issues)
    _truncate_field(data, "supplier", 128, issues)
    _truncate_field(data, "location", 128, issues)
    _truncate_field(data, "owner", 64, issues)
    _truncate_field(data, "department", 64, issues)
    _truncate_field(data, "status", 32, issues)
    _truncate_field(data, "remark", 255, issues)

    return data, issues


def _truncate_field(
    data: Dict[str, Any],
    field: str,
    max_len: int,
    issues: List[Dict[str, str]],
) -> None:
    v = data.get(field)
    if isinstance(v, str) and len(v) > max_len:
        data[field] = v[:max_len]
        issues.append(
            {
                "level": "warning",
                "field": field,
                "message": f"{field} 超过 {max_len} 字，已自动截断",
            }
        )


def get_field_dict() -> Dict[str, str]:
    """对外暴露系统字段字典（供前端列映射 select 使用）。"""
    return dict(SYSTEM_FIELD_DICT)


# ============ AI 一键补全（基于前置资产样本） ============

# 序列号 / 物理唯一标识类字段不应由 AI 编造
_NEVER_LLM_FILL = {"asset_code", "serial_number"}


def asset_to_reference(asset: Any) -> Dict[str, Any]:
    """把 models.Asset ORM 对象抽成 LLM 参考用的轻量字典。

    只保留对推断风格有帮助的字段，过滤掉空值减少 token。
    """
    fields = [
        "asset_class", "category", "brand", "model",
        "specification", "supplier", "location", "owner",
        "department", "status",
    ]
    ref: Dict[str, Any] = {}
    for f in fields:
        v = getattr(asset, f, None)
        if v is None:
            continue
        if isinstance(v, str):
            v = v.strip()
            if not v:
                continue
        ref[f] = v
    return ref


def _missing_fields_of(data: Dict[str, Any]) -> List[str]:
    """返回 data 中目前为空（None / 空串）且可由 LLM 补全的字段列表。"""
    missing: List[str] = []
    for k in ALLOWED_FIELD_KEYS:
        if k in _NEVER_LLM_FILL:
            continue
        v = data.get(k)
        if v is None or (isinstance(v, str) and not v.strip()):
            missing.append(k)
    return missing


def apply_ai_fill(
    rows: List[Dict[str, Any]],
    reference_assets: List[Dict[str, Any]],
    only_missing: bool = True,
) -> Tuple[List[Dict[str, Any]], int, int]:
    """对一批行调用 LLM 进行字段补全，并重新规范化 + 校验。

    参数：
        rows：list of {"row_index": int, "data": {...}}
        reference_assets：前置资产样本（已 asset_to_reference 化）
        only_missing：True 时只允许 LLM 覆盖空字段，已有非空字段保持不动

    返回：(new_rows, error_count, warning_count)
        new_rows 每行结构同 parse_workbook 输出：{"row_index", "data", "issues"}

    任何 LLM 异常都会抛 LLMUnavailable，由路由层捕获后转 502。
    """
    if not rows:
        return [], 0, 0

    payload_rows: List[Dict[str, Any]] = []
    for r in rows:
        data = dict(r.get("data") or {})
        payload_rows.append(
            {
                "row_index": r.get("row_index"),
                "data": data,
                "missing": _missing_fields_of(data),
            }
        )

    suggestions = llm_fill_missing_fields(payload_rows, reference_assets)

    # 用 row_index 做配对（容忍 LLM 偶尔返回顺序错乱）
    by_idx: Dict[Any, Dict[str, Any]] = {}
    for s in suggestions:
        if isinstance(s.get("row_index"), int):
            by_idx[s["row_index"]] = s.get("data") or {}

    new_rows: List[Dict[str, Any]] = []
    err_count = 0
    warn_count = 0
    for r in rows:
        ri = r.get("row_index")
        original = dict(r.get("data") or {})
        suggestion = by_idx.get(ri) or {}

        merged: Dict[str, Any] = dict(original)
        for k, v in suggestion.items():
            if k in _NEVER_LLM_FILL:
                continue
            if v is None:
                continue
            if isinstance(v, str) and not v.strip():
                continue
            if only_missing:
                cur = merged.get(k)
                is_empty = cur is None or (isinstance(cur, str) and not cur.strip())
                if not is_empty:
                    continue
            merged[k] = v

        # 重新跑一次规范化以触发类型转换 + 必填校验
        pseudo_mapping = {k: k for k in merged.keys()}
        data, issues = normalize_row(merged, pseudo_mapping)

        if any(i["level"] == "error" for i in issues):
            err_count += 1
        if any(i["level"] == "warning" for i in issues):
            warn_count += 1

        new_rows.append({"row_index": ri, "data": data, "issues": issues})

    return new_rows, err_count, warn_count


def preview_top_rows(
    file_bytes: bytes, filename: str, n: int = 10
) -> List[List[Optional[str]]]:
    """返回文件顶部最多 n 行的字符串化预览，给前端供用户切换表头行使用。"""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        raw = _read_csv_rows(file_bytes, n)
    else:
        raw = _read_xlsx_rows(file_bytes, n)

    preview: List[List[Optional[str]]] = []
    for row in raw[:n]:
        line: List[Optional[str]] = []
        for v in row:
            if v is None:
                line.append(None)
            elif isinstance(v, datetime):
                line.append(v.strftime("%Y-%m-%d"))
            elif isinstance(v, date):
                line.append(v.strftime("%Y-%m-%d"))
            else:
                s = str(v).strip()
                # 单元格内容过长时截断，避免响应过大
                line.append(s[:80] if s else None)
        preview.append(line)
    return preview
